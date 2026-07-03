"""
Report generation — exports ERP data to Excel (.xlsx) using openpyxl,
or a simple text summary for PDF-like output via reportlab.
"""

from typing import Any, Dict
from pathlib import Path
from datetime import datetime

from database.connection import db
from actions.registry import register
from config import REPORT_OUTPUT_DIR


@register("GENERATE_REPORT")
async def handle_generate_report(entities: Dict[str, Any], user_id: int) -> str:
    products = entities.get("products", [])
    numbers = entities.get("numbers", [])

    if products:
        filename = _export_product_stock_report(products[0])
        if filename:
            return f"📄 تم تصدير التقرير: {filename}"
        return "لم يتم العثور على بيانات لهذا المنتج."

    filename = _export_general_report()
    if filename:
        return f"📄 تم تصدير التقرير العام للمخزون: {filename}"
    return "لم يتم تصدير التقرير."


def _export_product_stock_report(product_name: str) -> str | None:
    product = db.fetch_one(
        "SELECT id, name, unit FROM products WHERE name = %s",
        (product_name,),
    )
    if not product:
        return None

    stocks = db.fetch_all(
        "SELECT w.name AS warehouse, COALESCE(s.quantity, 0) AS qty "
        "FROM warehouses w "
        "LEFT JOIN stock s ON s.warehouse_id = w.id AND s.product_id = %s "
        "ORDER BY w.name",
        (product["id"],),
    )

    movements = db.fetch_all(
        """
        SELECT type, quantity, date, notes
        FROM stock_movements
        WHERE product_id = %s
        ORDER BY date DESC LIMIT 50
        """,
        (product["id"],),
    )

    from openpyxl import Workbook

    wb = Workbook()

    # Sheet 1: Stock levels
    ws1 = wb.active
    ws1.title = "المخزون"
    ws1.append(["المستودع", "الكمية"])
    for s in stocks:
        ws1.append([s["warehouse"], float(s["qty"])])

    # Sheet 2: Movements
    ws2 = wb.create_sheet("الحركات")
    ws2.append(["النوع", "الكمية", "التاريخ", "ملاحظات"])
    for m in movements:
        ws2.append([m["type"], float(m["quantity"]), str(m["date"] or ""), m.get("notes", "")])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_{product['name']}_{timestamp}.xlsx"
    filepath = REPORT_OUTPUT_DIR / filename
    wb.save(str(filepath))
    return str(filepath)


def _export_general_report() -> str | None:
    products = db.fetch_all(
        """
        SELECT p.id, p.name, p.type, p.unit, p.min_stock,
               COALESCE(SUM(s.quantity), 0) AS total_qty
        FROM products p
        LEFT JOIN stock s ON s.product_id = p.id
        GROUP BY p.id, p.name, p.type, p.unit, p.min_stock
        ORDER BY p.name
        """
    )
    if not products:
        return None

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير المخزون"
    ws.append(["الكود", "الاسم", "النوع", "الوحدة", "الكمية", "حد الأمان"])
    for p in products:
        ws.append([
            p["id"],
            p["name"],
            p["type"],
            p.get("unit", ""),
            float(p["total_qty"]),
            float(p["min_stock"]) if p["min_stock"] else "",
        ])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inventory_report_{timestamp}.xlsx"
    filepath = REPORT_OUTPUT_DIR / filename
    wb.save(str(filepath))
    return str(filepath)
