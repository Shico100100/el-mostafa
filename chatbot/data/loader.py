"""
Document loader — scans chatbot/data/ and extracts text from
PDF, DOCX, XLSX, TXT, CSV, and PPTX files.

Usage:
    from data.loader import load_documents, search_documents
    docs = load_documents()          # returns [{name, path, text, pages}, ...]
    hits = search_documents("PVC")   # keyword search across loaded docs
"""

import os
import re
import sys
from typing import Any, Dict, List

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)))

_documents: List[Dict[str, Any]] = []
_loaded = False


def _extract_text_from_pdf(path: str) -> str:
    from pypdf import PdfReader
    reader = PdfReader(path)
    parts = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            parts.append(t)
    return "\n".join(parts)


def _extract_text_from_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_text_from_xlsx(path: str) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            parts.append(f"[{sheet}]\n" + "\n".join(rows))
    return "\n\n".join(parts)


def _extract_text_from_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def _extract_text_from_csv(path: str) -> str:
    import csv
    rows = []
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(" | ".join(row))
    return "\n".join(rows)


EXTRACTORS = {
    ".pdf": _extract_text_from_pdf,
    ".docx": _extract_text_from_docx,
    ".doc": _extract_text_from_docx,
    ".xlsx": _extract_text_from_xlsx,
    ".xls": _extract_text_from_xlsx,
    ".txt": _extract_text_from_txt,
    ".csv": _extract_text_from_csv,
}


def load_documents() -> List[Dict[str, Any]]:
    global _documents, _loaded
    _documents = []
    _loaded = False

    if not os.path.isdir(DATA_DIR):
        return _documents

    for fname in os.listdir(DATA_DIR):
        fpath = os.path.join(DATA_DIR, fname)
        if not os.path.isfile(fpath):
            continue
        ext = os.path.splitext(fname)[1].lower()
        if ext not in EXTRACTORS:
            continue

        try:
            text = EXTRACTORS[ext](fpath)
        except Exception as e:
            print(f"[Docs] Failed to read {fname}: {e}", file=sys.stderr)
            continue

        if text.strip():
            _documents.append({
                "name": fname,
                "path": fpath,
                "text": text,
                "size": len(text),
            })

    _loaded = True
    print(f"[Docs] Loaded {len(_documents)} documents from data/", flush=True)
    return _documents


def reload_documents() -> List[Dict[str, Any]]:
    return load_documents()


def search_documents(query: str, max_results: int = 3) -> List[Dict[str, Any]]:
    if not _loaded:
        load_documents()

    q = query.lower()
    results = []
    for doc in _documents:
        text_lower = doc["text"].lower()
        score = text_lower.count(q)
        if score > 0:
            results.append({**doc, "score": score})

    results.sort(key=lambda x: x["score"], reverse=True)
    return results[:max_results]


def get_document(name: str) -> Dict[str, Any] | None:
    if not _loaded:
        load_documents()
    for doc in _documents:
        if doc["name"] == name:
            return doc
    return None


def list_documents() -> List[Dict[str, Any]]:
    if not _loaded:
        load_documents()
    return [{"name": d["name"], "size": d["size"]} for d in _documents]
